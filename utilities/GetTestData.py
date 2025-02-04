import os
from pathlib import Path

import openpyxl


class GetTestData:

    @staticmethod
    def get_test_data(test_case_name):
        test_data = {}
        test_data_list = []
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.abspath(os.path.join(script_dir, ".."))
        excel_path = Path(project_root) / "data" / "TestData.xlsx"
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    test_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                test_data_list.append(test_data.copy())
        return test_data_list
